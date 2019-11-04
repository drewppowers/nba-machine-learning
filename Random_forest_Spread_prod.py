import pandas as pd
from sportsreference.nba.teams import Teams
import sportsreference
from sportsreference.nba.schedule import Schedule
import datetime
from datetime import timedelta, date
import pickle
import os
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import StandardScaler
import pandas
from openpyxl import load_workbook

yest = datetime.date.today() - datetime.timedelta(1)
today = datetime.date.today()

def get_stats(team_abbr, team_stats):
    pd.options.mode.chained_assignment = None  # default='warn'
    sched = Schedule(team_abbr)
    sched_df = sched.dataframe
    sched_df['datetime'] = sched_df['datetime'].dt.date
    team_df = pd.DataFrame()
    for game in sched:
        team_df = team_df.append(game.dataframe)
    team_df['datetime'] = team_df['datetime'].dt.date
    curr_sched = team_df[team_df['datetime'] <= yest]
    curr_sched['stk'] = curr_sched['streak'].str.replace('L ', '-').str.replace('W ', '')
    last_game = curr_sched.iloc[-1]
    # 'lastTen'
    lTen = curr_sched.tail(10)
    if 'Win' in lTen['result'].value_counts().index:
        last_game['lastTen'] = lTen['result'].value_counts()['Win']
    else:
        last_game['lastTen'] = 0
    # 'winPercent'
    last_game = last_game.fillna(0)
    last_game['winPercent'] = last_game['wins']/last_game['game']
    last_game = last_game.fillna(0)
    # Home win
    if 'Win' in curr_sched[curr_sched['location']=='Home']['result'].value_counts().index:
        last_game['homeWin'] = curr_sched[curr_sched['location']=='Home']['result'].value_counts()['Win']
    else:
        last_game['homeWin'] = 0
    # Home loss
    if 'Loss' in curr_sched[curr_sched['location']=='Home']['result'].value_counts().index:
        last_game['homeLoss'] = curr_sched[curr_sched['location']=='Home']['result'].value_counts()['Loss']
    else:
        last_game['homeLoss'] = 0
    # Away win
    if 'Win' in curr_sched[curr_sched['location']=='Away']['result'].value_counts().index:
        last_game['awayWin'] = curr_sched[curr_sched['location']=='Away']['result'].value_counts()['Win']
    else:
        last_game['awayWin'] = 0
    # Away loss
    if 'Loss' in curr_sched[curr_sched['location']=='Away']['result'].value_counts().index:
        last_game['awayLoss'] = curr_sched[curr_sched['location']=='Away']['result'].value_counts()['Loss']
    else:
        last_game['awayLoss'] = 0
    last_game['homeWin_Percent'] = last_game['homeWin']/(last_game['homeWin'] + last_game['homeLoss'])
    last_game['awayWin_Percent'] = last_game['awayWin']/(last_game['awayWin'] + last_game['awayLoss'])
    last_game = last_game[['stk', 'lastTen', 'winPercent', 'homeWin_Percent', 'awayWin_Percent']]
     # 'teamPTS_pg', 
    team_stats['teamPTS_pg'] = team_stats['points']/team_stats['games_played']
    # 'teamTO_pg', 
    team_stats['teamTO_pg'] = team_stats['turnovers']/team_stats['games_played']
    # 'teamFG%_pg', 
    team_stats['teamFG%_pg'] = team_stats['field_goal_percentage']
    # 'team2P%_pg',
    team_stats['team2P%_pg'] = team_stats['two_point_field_goal_percentage']
    # 'team3PA_pg', 
    team_stats['team3PA_pg'] = team_stats['three_point_field_goal_attempts']/team_stats['games_played']
    # 'teamTRB_perc', 
    team_stats['teamTRB_perc'] = team_stats['total_rebounds'] * 100 /(team_stats['total_rebounds'] + team_stats['opp_total_rebounds'])
    # 'teamTO_perc', 
    team_stats['teamTO_perc'] = team_stats['turnovers']*100/(team_stats['field_goal_attempts']*.44+team_stats['free_throw_attempts'] + team_stats['turnovers'])
    # 'PPS',
    team_stats['PPS'] = team_stats['points']/team_stats['field_goal_attempts']
    # 'teamSTL/TO', 
    team_stats['teamSTL/TO'] = team_stats['steals']/team_stats['turnovers']
    # pyth%13.91', 
    # Calculation: ptsFor ^ 13.91 / (ptsFor ^ 13.91 + ptsAgnst ^ 13.91)
    team_stats['pyth%13.91'] = team_stats['points']**13.91 / (team_stats['points']**13.91+team_stats['opp_points']**13.91)
    # 'lpyth13.91',
    team_stats['lpyth13.91'] = 82 - team_stats['pyth%13.91'] * 82
    #mov
    team_stats['mov'] = (team_stats['points']-team_stats['opp_points'])/team_stats['games_played']
    stats = team_stats.loc[team_abbr].append(last_game)
    stats = stats[['teamPTS_pg','teamTO_pg','teamFG%_pg','team2P%_pg','team3PA_pg', 'teamTRB_perc',
                   'PPS','teamSTL/TO','pyth%13.91','lpyth13.91','teamTO_perc','mov', 'stk', 'lastTen',
                   'winPercent','homeWin_Percent', 'awayWin_Percent'] ]
    return stats

def games_today(team_abbr):
    sched = Schedule(team_abbr)
    sched_df = sched.dataframe
    sched_df['team_abbr'] = team_abbr
    sched_df = sched_df[['datetime', 'team_abbr', 'opponent_abbr', 'location', 
        'time']]
    sched_df = sched_df.reset_index(drop = True)
    sched_df = sched_df[sched_df['datetime']==today]
    return sched_df

team_stats = pd.DataFrame()
for team in Teams():
    team_stats = team_stats.append(team.dataframe)
teams = Teams()
team_list = []
for team in teams:
    team_list.append(team.abbreviation)
nba = pd.DataFrame()
game_today = pd.DataFrame()
for i in range(len(team_list)):
    team_abbr = team_list[i]
    data = get_stats(team_abbr, team_stats)
    data['teamAbbr'] = team_abbr
    nba = nba.append(data, ignore_index = True)
nba = nba.set_index('teamAbbr')
matchups = pd.DataFrame()
for i in range(len(team_list)):
    
    matchups = matchups.append(games_today(team_list[i]))

matchups = matchups[matchups['location'] == 'Home'].sort_values('time')
matchups = matchups.set_index('team_abbr')
matchups = matchups.merge(nba.drop('awayWin_Percent', axis = 1).add_suffix('_1'), how = 'left',  
               left_index=True, right_index = True)
matchups = matchups.merge(nba.drop('homeWin_Percent', axis = 1).add_suffix('_2'), how = 'left',  
               left_on = 'opponent_abbr', right_index = True)
matchups = matchups.rename(columns={'homeWin_Percent_1': 'homeAway_winPercent_1', 
                                    'awayWin_Percent_2': 'homeAway_winPercent_2'})

os.chdir('G://My Drive//NBA Machine Learning//Code')

# import model. its in nba machine learning/code/spread
# load the model from disk
clf = pickle.load(open('./Spread/nba_11_3_random_forest_spread_model.sav', 'rb'))

df = pd.read_csv('../Code/11_1_nba_line_trimmed.csv')

df = df.drop(['Unnamed: 0', 'gmDate', 'team1Score', 'team2Score', 'teamAbbr_1', 'teamAbbr_2', 
             'total1', 'over_price', 'under_price', 'over_hit',
       'spread2', 'spread1', 'spread_price2', 'spread_price1', 'ml2', 'ml1',
       'actual_total', 'actual_spread'], axis = 1)

matchups = matchups[['teamPTS_pg_1', 'teamTO_pg_1', 'teamFG%_pg_1', 'team2P%_pg_1',
       'team3PA_pg_1', 'teamTRB_perc_1', 'teamTO_perc_1', 'PPS_1',
       'teamSTL/TO_1', 'stk_1', 'lastTen_1', 'pyth%13.91_1', 'lpyth13.91_1',
       'winPercent_1', 'homeAway_winPercent_1', 'mov_1', 'teamPTS_pg_2',
       'teamTO_pg_2', 'teamFG%_pg_2', 'team2P%_pg_2', 'team3PA_pg_2',
       'teamTRB_perc_2', 'teamTO_perc_2', 'PPS_2', 'teamSTL/TO_2', 'stk_2',
       'lastTen_2', 'pyth%13.91_2', 'lpyth13.91_2', 'winPercent_2',
       'homeAway_winPercent_2', 'mov_2']]

X = df[matchups.columns]

y = df['spread_cover']


X = X.append(matchups)

scaled_features = StandardScaler().fit_transform(X.values)

scaled_features_df = pd.DataFrame(scaled_features, index=X.index, columns=X.columns)

new_games = scaled_features_df.loc[matchups.index]

matchups['Cover_Spread_pred'] = clf.predict(new_games)

bets = pd.DataFrame()
for i in range(len(matchups.index)):
    bets = bets.append(games_today(matchups.index[i]))
bets = bets.set_index('team_abbr')

bets = bets.merge(matchups['Cover_Spread_pred'], left_index=True, right_index=True)

bets = pd.DataFrame()
for i in range(len(matchups.index)):
    bets = bets.append(games_today(matchups.index[i]))
bets = bets.set_index('team_abbr')

book = load_workbook('NBA_Random_Forest_Spread_History.xlsx')
writer = pandas.ExcelWriter('NBA_Random_Forest_Spread_History.xlsx', engine='openpyxl')
writer.book = book
writer.sheets = {ws.title: ws for ws in book.worksheets}

for sheetname in writer.sheets:
    bets.to_excel(writer,sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row, index = True,header= False)

writer.save()

matchups['Cover_Spread']